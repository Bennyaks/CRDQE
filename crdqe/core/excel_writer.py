"""
===========================================================
Writes output Excel report.
===========================================================
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:

    def __init__(self, settings):

        output = Path(settings["output"]["folder"])

        self.report_path = (
            output /
            "CRDQE Validation Report.xlsx"
        )

    # -----------------------------------------------------
    # Prepare cleaned dataframe
    # -----------------------------------------------------

    def write_cleaned(self, dataframe):

        df = dataframe.copy()

        date_columns = [
            c for c in df.columns
            if "date" in c.lower()
        ]

        for col in date_columns:

            df[col] = (
                pd.to_datetime(
                    df[col],
                    errors="coerce",
                    dayfirst=True
                )
                .dt.strftime("%d/%m/%Y")
                .fillna("")
            )

        if "entry_number" in df.columns:

            df["entry_number"] = (
                pd.to_numeric(
                    df["entry_number"],
                    errors="coerce"
                )
                .astype("Int64")
            )

        return df

    # -----------------------------------------------------
    # Prepare Flag dataframe
    # -----------------------------------------------------

    def write_flags(self, dataframe):

        return dataframe.copy()

    # -----------------------------------------------------
    # Write complete report
    # -----------------------------------------------------

    def write_report(
        self,
        cleaned_df,
        flags_df,
        report
    ):

        cleaned_df = self.write_cleaned(cleaned_df)
        flags_df = self.write_flags(flags_df)

        summary_df = pd.DataFrame({

            "Metric": [

                "Rows",

                "Columns",

                "Issues Found",

                "Quality Score (%)",

                "Current Registrations",

                "Late Registrations",
                "Health Facility Date Corrections"

            ],

            "Value": [

                report["rows"],

                report["columns"],

                report["issues"],

                report["quality_score"],

                report.get("current_cases", 0),

                report.get("late_cases", 0),
                report.get("swapped_dates", 0)
            ]

        })

        issues_df = pd.DataFrame(

            list(report["issues_by_field"].items()),

            columns=[

                "Field",

                "Issues"

            ]

        )

        with pd.ExcelWriter(
            self.report_path,
            engine="openpyxl"
        ) as writer:

            summary_df.to_excel(
                writer,
                sheet_name="Validation Summary",
                index=False,
                startrow=0
            )

            issues_df.to_excel(
                writer,
                sheet_name="Validation Summary",
                startrow=len(summary_df) + 3,
                index=False
            )

            cleaned_df.to_excel(
                writer,
                sheet_name="Cleaned Data",
                index=False
            )

            flags_df.to_excel(
                writer,
                sheet_name="Flag Report",
                index=False
            )

        # VERY IMPORTANT
        # Workbook must be closed before formatting

        self._format(self.report_path)

    # -----------------------------------------------------
    # Formatting
    # -----------------------------------------------------

    def _format(self, path):

        workbook = load_workbook(path)

        header_fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        header_font = Font(

            bold=True,

            color="FFFFFF"

        )

        thin = Side(style="thin")

        border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )

        # Format every worksheet

        for worksheet in workbook.worksheets:

            for cell in worksheet[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            worksheet.freeze_panes = "A2"

            worksheet.auto_filter.ref = worksheet.dimensions

            for column in worksheet.columns:

                width = max(

                    len(str(cell.value))
                    if cell.value is not None
                    else 0

                    for cell in column

                )

                worksheet.column_dimensions[
                    get_column_letter(column[0].column)
                ].width = min(width + 4, 50)

        workbook.save(path)