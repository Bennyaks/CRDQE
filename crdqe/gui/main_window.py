"""
===========================================================
CRDQE Main Window
-----------------------------------------------------------
Top-level application window. Creates the theme, menu,
toolbar, sidebar, dashboard and status bar, and wires them
together with the validation engine.
===========================================================
"""

import os
import threading
import tkinter as tk

from tkinter import filedialog

from openpyxl import load_workbook

from crdqe.utils.paths import BASE_DIR, ASSETS_DIR
from crdqe.core.config_manager import ConfigManager
from crdqe.core.engine import CRDQEEngine

from crdqe.gui.dashboard import Dashboard
from crdqe.gui.dialogs import Dialogs
from crdqe.gui.menu import MenuBar
from crdqe.gui.sidebar import Sidebar
from crdqe.gui.status_bar import StatusBar
from crdqe.gui.splash import SplashScreen
from crdqe.gui.theme import Theme
from crdqe.gui.toolbar import Toolbar
from crdqe.gui.widgets import (
    LogConsole,
    ProgressFrame,
)


OUTPUT_FOLDER = "output"


class MainWindow(tk.Tk):

    def __init__(self):

        super().__init__()
        self.withdraw()

        SplashScreen(self)

        self.after(
            2500,
            self.deiconify
        )

        self.initialize_window()

        self.initialize_state()

        self.create_components()

        self.create_layout()

        self.connect_signals()
        try:
            self.iconbitmap(str(ASSETS_DIR / "logo.ico"))
        except Exception as exc:
            print(f"Could not load application icon: {exc}")
        self.update()

    


    def initialize_window(self):

        self.title("CRDQE - Civil Registration Data Quality Engine")

        self.geometry("1200x750")

        self.minsize(1000, 650)

        Theme.configure()

    def initialize_state(self):
        self.validation_running = False

        self.workbook_path = None

        self.engine = CRDQEEngine()

        self.output_folder = self._resolve_output_folder()

    def _resolve_output_folder(self):
        """
        Reads the real output folder from config/settings.yaml
        (settings["output"]["folder"]) -- this is the same path
        FileManager/ExcelWriter actually write reports to, so the
        "Open Output Folder" button opens the right place instead
        of a hardcoded guess. Anchored to BASE_DIR (crdqe.utils.paths)
        so it matches file_manager.py / excel_writer.py exactly,
        regardless of the working directory the app was launched from.
        """

        try:

            settings = ConfigManager().load()

            return str(BASE_DIR / settings["output"]["folder"])

        except Exception as exc:

            print(f"Could not resolve output folder from config: {exc}")

            return str(BASE_DIR / OUTPUT_FOLDER)

    # ------------------------------------------------------
    # Create_components
    # ------------------------------------------------------

    def create_components(self):

        self.menu = MenuBar(
            self,
            open_callback=self.browse_workbook,
            validate_callback=self.run_validation,
            output_folder=self.output_folder
        )

        self.toolbar = Toolbar(self)

        self.status_bar = StatusBar(self)

    # ---- Create_Layout--------------------------------
    def create_layout(self):

        body = tk.Frame(
            self,
            bg=Theme.BACKGROUND
        )

        body.pack(
            fill="both",
            expand=True
        )

        body.columnconfigure(1, weight=1)

        body.rowconfigure(0, weight=1)

        self.sidebar = Sidebar(
            body,
            browse_callback=self.browse_workbook,
            worksheet_callback=self.on_worksheet_changed
        )
        self.sidebar.grid(
            row=0,
            column=0,
            sticky="ns",
            padx=(10,5),
            pady=10
        )

        content = tk.Frame(
            body,
            bg=Theme.BACKGROUND
        )

        content.grid(
            row=0,
            column=1,
            sticky="nsew"
        )

        content.columnconfigure(0, weight=1)

        content.rowconfigure(2, weight=1)

        self.dashboard = Dashboard(content)

        self.dashboard.grid(
            row=0,
            column=0,
            sticky="ew"
        )

        self.progress = ProgressFrame(content)

        self.progress.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=20,
            pady=(10,0)
        )

        self.log = LogConsole(content)

        self.log.grid(
            row=2,
            column=0,
            sticky="nsew",
            padx=20,
            pady=(10,20)
        )
        self.toolbar.run_button.configure(
            state="disabled"
        )
    # --- connect_signal -----------------------------------
    def connect_signals(self):
        """
        Connect widget events.

        Reserved for future event bindings.
        """
        pass

    # ------------------------------------------------------
    # ACTIONS
    # ------------------------------------------------------

    def browse_workbook(self):

        path = filedialog.askopenfilename(
            title="Select Workbook",
            filetypes=[("Excel Workbook", "*.xlsx *.xlsm")]
        )

        if not path:
            return

        self.workbook_path = path

        try:

            workbook = load_workbook(
                path,
                read_only=True
            )

            self.sidebar.set_worksheets(
                workbook.sheetnames
            )
            if workbook.sheetnames:

                self.sidebar.worksheet_combo.current(0)

                self.on_worksheet_changed()

        finally:

            workbook.close()
        self.sidebar.set_workbook(
            os.path.basename(path)
        )

        self.status_bar.set_dataset(os.path.basename(path))
        self.status_bar.set_status("Workbook loaded")
        self.log.append(f"Loaded workbook: {path}")
        self.toolbar.run_button.configure(
            state="normal"
        )
    def on_worksheet_changed(self):
        """
        Called whenever the user selects a different worksheet.
        """

        worksheet = self.sidebar.get_selected_worksheet()

        if not worksheet:
            return

        self.status_bar.set_status(
            f"Worksheet selected: {worksheet}"
        )

        self.detect_dataset()

    def detect_dataset(self):
        """
        Detect whether the selected worksheet
        contains Birth or Death registrations.
        """

        try:
            # Fallback implementation using openpyxl to avoid
            # dependency on ExcelReader/HeaderDetector/etc.
            wb = load_workbook(self.workbook_path, read_only=True)
            ws = wb[self.sidebar.get_selected_worksheet()]

            # Read first non-empty row as header
            header = None
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if row and any(cell is not None for cell in row):
                    header = [str(cell).strip() if cell is not None else "" for cell in row]
                    break

            wb.close()

            if not header:
                self.sidebar.dataset_combo.set("Auto Detect")
                return

            cols = [c.lower() for c in header]

            # Simple detection heuristics
            if any("birth" in c or "dob" in c for c in cols):
                dataset = "Birth"
            elif any("death" in c or "dod" in c for c in cols):
                dataset = "Death"
            else:
                dataset = "Auto Detect"

            self.sidebar.dataset_combo.set(dataset)

            self.status_bar.set_status(f"Detected dataset: {dataset}")

        except Exception:
            self.sidebar.dataset_combo.set("Auto Detect")

    def run_validation(self):

        if not self.workbook_path:
            Dialogs.warning(
                "No Workbook",
                "Please open a workbook before running validation."
            )
            return

        worksheet = self.sidebar.get_selected_worksheet()

        if not worksheet:
            Dialogs.warning(
                "Worksheet",
                "Please select a worksheet."
            )
            return

        month = self.sidebar.get_selected_month()

        if not month:
            Dialogs.warning(
                "Registration Month",
                "Please select the registration month."
            )
            return

        self.status_bar.set_status("Validating...")
        self.progress.start()
        self.toolbar.run_button.configure(state="disabled")

        if self.validation_running:
            return

        self.validation_running = True

        thread = threading.Thread(
            target=self._run_validation_thread,
            args=(worksheet, month),
            daemon=True
        )

        thread.start()

    def _run_validation_thread(self, worksheet, month):

        try:
            print("Worksheet:", self.sidebar.get_selected_worksheet())
            print("Dataset :", self.sidebar.get_selected_dataset())

            result = self.engine.run(
                workbook_path=self.workbook_path,
                worksheet=worksheet,
                registration_month=month,
                callback=lambda msg: self.after(0, self.log.append, msg),
            )

            self.after(0, self._on_validation_complete, result)

        except Exception as exc:
            self.after(0, self._on_validation_error, exc)

    def _on_validation_complete(self, result):

        self.progress.stop()
        self.toolbar.run_button.configure(state="normal")
        self.validation_running = False

        # CRDQEEngine.run() returns False when the workbook wasn't found,
        # and a dict on success -- handle the failure case first.
        if not result:

            self.status_bar.set_status("Validation failed")
            self.log.append("Workbook not found or validation could not run.")
            Dialogs.error(
                "Validation Failed",
                "Workbook not found or validation could not run."
            )
            return

        self.last_result = result
        self.status_bar.set_status("Validation complete")
        self.log.append(
            "Validation completed successfully."
        )

        records = result.get("records", 0)
        issues = result.get("issues", 0)
        current = result.get("current", 0)
        late = result.get("late", 0)


        self.status_bar.set_records(records)
        self.status_bar.set_issues(issues)
        self.dashboard.update_statistics(records, issues, current, late)
        quality = (
            "100%"
            if records == 0
            else f"{max(0, round((1 - issues / records) * 100))}%"
        )

        self.sidebar.update_statistics(
            records,
            issues,
            quality
        )

        Dialogs.info(
            "Validation Complete",
            f"Processed {records} records, found {issues} issues."
        )

    def _on_validation_error(self, exc):

        self.progress.stop()
        self.toolbar.run_button.configure(state="normal")
        self.validation_running = False
        self.status_bar.set_status("Validation failed")

        self.log.append(f"ERROR: {exc}")

        Dialogs.error("Validation Error", str(exc))

    def open_output_folder(self):
        """
        Opens the folder containing all generated outputs.
        """

        folder = self.output_folder

        print("Opening:", folder)

        if not os.path.isdir(folder):
            Dialogs.warning(
                "Output Folder",
                "No output folder has been generated yet."
            )
            return

        os.startfile(folder)
    
    def clear_log(self):

        self.log.clear()